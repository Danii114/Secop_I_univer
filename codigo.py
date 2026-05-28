import os
import re
import sys
import ast
import time
import smtplib
import requests
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, date
from sentence_transformers import SentenceTransformer, util
from rank_bm25 import BM25Okapi
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import torch

# =========================================================
# 🔹 CONFIG
# =========================================================
EMAIL_SENDER = os.getenv("EMAIL_SENDER")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")

EMAIL_TO = [
    "danimorav05@gmail.com",
]

URL           = "https://www.datos.gov.co/resource/f789-7hwg.json"
ARCHIVO_EXCEL = f"alertas_secop_{date.today()}.xlsx"
DIAS_ATRAS    = 5

SEMANTICA_TOP_K     = 0     
SEMANTICA_MIN_SCORE = 0.60
BM25_MIN_SCORE      = 4.0
# =========================================================
# 🔹 QUERIES SEMÁNTICAS
# =========================================================
QUERIES = [
    # Educación
    "fortalecimiento academico pruebas saber estudiantes",
    "capacitacion docentes instituciones educativas",
    "acompanamiento pedagogico educacion media",
    "competencias academicas pruebas icfes",

    # Educación superior
    "convenio educacion superior universidad",
    "convenio universidad institucion educativa",
    "programa academico universidad",

    # Tecnología educativa
    "software educativo plataforma academica",
    "licenciamiento software educativo",
    "sistema informacion academico",

    # Investigación
    "investigacion cientifica universidad",
    "ciencia tecnologia innovacion universidad",
    "transferencia tecnologica academia",

    # Extensión
    "bienestar universitario estudiantes",
    "practicas academicas universitarias",
]

# =========================================================
# 🔹 KEYWORDS BM25
# =========================================================
KEYWORDS_BM25 = [

    # Educación
    "fortalecimiento academico",
    "pruebas saber",
    "simulacros icfes",
    "acompanamiento pedagogico",
    "capacitacion docentes",
    "competencias academicas",

    # Educación superior
    "educacion superior",
    "universidad",
    "universitaria",
    "institucion educativa",

    # Tecnología educativa
    "software educativo",
    "plataforma academica",
    "sistema academico",
    "licenciamiento software",

    # Investigación
    "investigacion cientifica",
    "ciencia tecnologia innovacion",
    "transferencia tecnologica",

    # Extensión
    "practicas academicas",
    "bienestar universitario",

]
# =========================================================
# 🔹 HELPERS
# =========================================================

def limpiar_texto(texto):
    texto = str(texto).lower()
    reemplazos = str.maketrans("áéíóúàèìòùäëïöüñ", "aeiouaeiouaeioun")
    texto = texto.translate(reemplazos)
    texto = re.sub(r"[^a-z0-9\s]", " ", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto

def safe_float(val):
    try:
        return float(pd.to_numeric(val, errors="coerce") or 0)
    except Exception:
        return 0.0

def extraer_url(val):
    if isinstance(val, dict):
        return val.get("url") or val.get("href") or str(list(val.values())[0])
    if isinstance(val, str):
        val_strip = val.strip()
        if val_strip.startswith("{"):
            try:
                d = ast.literal_eval(val_strip)
                return d.get("url") or d.get("href") or str(list(d.values())[0])
            except:
                pass
        return val
    return str(val) if val else ""

def limpiar_valor(v):
    if isinstance(v, dict):
        return str(list(v.values())[0])
    if isinstance(v, list):
        return str(v[0]) if v else ""
    return v

def enviar_aviso_sin_datos(motivo):
    print(f"⚠️  {motivo} — enviando aviso por correo")
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"ℹ️ SECOP sin alertas · {date.today()}"
        msg["From"]    = EMAIL_SENDER
        msg["To"]      = ", ".join(EMAIL_TO)
        cuerpo = f"<html><body><p>Pipeline ejecutado sin resultados.<br><b>Motivo:</b> {motivo}</p></body></html>"
        msg.attach(MIMEText(cuerpo, "html"))
        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.sendmail(EMAIL_SENDER, EMAIL_TO, msg.as_string())
        print("✅ Aviso enviado")
    except Exception as e:
        print(f"❌ No se pudo enviar aviso: {e}")
    sys.exit(0)

# =========================================================
# 🔹 PASO 1 — DESCARGA CON PAGINACIÓN
# =========================================================

print("=" * 60)
print(f"  🚀 PIPELINE SECOP I — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
print("=" * 60)

hoy        = datetime.now()
inicio     = (hoy - timedelta(days=DIAS_ATRAS)).replace(hour=0,  minute=0,  second=0)
fin        = (hoy - timedelta(days=1)).replace(         hour=23, minute=59, second=59)
inicio_str = inicio.strftime("%Y-%m-%dT%H:%M:%S")
fin_str    = fin.strftime("%Y-%m-%dT%H:%M:%S")
fecha_str  = fin.strftime("%Y-%m-%d")

print(f"\n🌐 Descargando SECOP I ({inicio.strftime('%Y-%m-%d')} → {fecha_str})...")

todos  = []
offset = 0
limit  = 5000

while True:
    params = {
        "$where": (
            f"fecha_de_cargue_en_el_secop between '{inicio_str}' and '{fin_str}' "
            f"AND estado_del_proceso = 'CONVOCADO'"
        ),
        "$limit":  limit,
        "$offset": offset,
    }
    try:
        resp  = requests.get(URL, params=params, timeout=(10, 60))
        resp.raise_for_status()
        batch = resp.json()
    except requests.exceptions.Timeout:
        print(f"⏱️  Timeout en offset {offset} — reintentando en 10s")
        time.sleep(10)
        continue
    except Exception as e:
        print(f"❌ Error descarga: {e}")
        break

    if not batch:
        break
    todos.extend(batch)
    print(f"   📦 {len(todos):,} registros...")
    if len(batch) < limit:
        break
    offset += limit
    time.sleep(0.3)

if not todos:
    enviar_aviso_sin_datos("La API no devolvió registros convocados en el período")

df = pd.DataFrame(todos)
print(f"✅ Descargados: {len(df):,} contratos CONVOCADOS")

# =========================================================
# 🔹 PASO 2 — TEXTO NLP
# =========================================================

print("\n🧠 Preparando texto...")

df["texto_busqueda"] = (
    df["detalle_del_objeto_a_contratar"]
    .fillna("")
    .apply(limpiar_texto)
)
df = df[df["texto_busqueda"].str.len() > 10].reset_index(drop=True)
print(f"✅ Textos válidos: {len(df):,}")

if len(df) == 0:
    enviar_aviso_sin_datos("Ningún convocado tenía texto válido")

# =========================================================
# 🔹 PASO 3 — BÚSQUEDA SEMÁNTICA
# =========================================================

print("\n🧠 Cargando modelo semántico...")
modelo = SentenceTransformer("paraphrase-multilingual-MiniLM-L12-v2")

print("⚙️  Generando embeddings...")
corpus_embeddings = modelo.encode(
    df["texto_busqueda"].tolist(),
    batch_size=64,
    show_progress_bar=True,
    convert_to_tensor=True,
)

print("\n🔍 Búsqueda semántica...")
indices_semantica = set()

for query in QUERIES:
    q_emb  = modelo.encode(query, convert_to_tensor=True)
    scores = util.cos_sim(q_emb, corpus_embeddings)[0]

    # ← SOLO por umbral, sin top_k
    hits = torch.where(scores >= SEMANTICA_MIN_SCORE)[0].tolist()
    indices_semantica.update(hits)
    print(f"   ✔ '{query[:55]}' → {len(hits)} hits")

print(f"✅ Semántica total: {len(indices_semantica):,}")

# =========================================================
# 🔹 PASO 4 — BM25
# =========================================================

print("\n🔍 BM25...")
corpus_tok   = [doc.split() for doc in df["texto_busqueda"]]
bm25         = BM25Okapi(corpus_tok)
indices_bm25 = set()

for keyword in KEYWORDS_BM25:
    scores = bm25.get_scores(limpiar_texto(keyword).split())
    hits   = np.where(scores >= BM25_MIN_SCORE)[0]
    indices_bm25.update(hits.tolist())
    print(f"   ✔ '{keyword}' → {len(hits)} hits")

print(f"✅ BM25 total: {len(indices_bm25):,}")

# =========================================================
# 🔹 PASO 5 — COMBINAR Y DEDUPLICAR
# =========================================================

print("\n🔗 Combinando resultados...")

indices_finales = indices_semantica | indices_bm25
print(f"   Solo semántica : {len(indices_semantica - indices_bm25):,}")
print(f"   Solo BM25      : {len(indices_bm25 - indices_semantica):,}")
print(f"   Ambos métodos  : {len(indices_semantica & indices_bm25):,}")
print(f"   Total únicos   : {len(indices_finales):,}")

if not indices_finales:
    enviar_aviso_sin_datos("Ningún proceso superó los umbrales de relevancia")

df_final = df.iloc[list(indices_finales)].copy()
df_final["origen"] = [
    "ambos"     if i in indices_semantica and i in indices_bm25
    else ("semantica" if i in indices_semantica else "bm25")
    for i in list(indices_finales)
]
df_final = df_final.reset_index(drop=True)
df_final["cuantia_proceso"] = pd.to_numeric(
    df_final.get("cuantia_proceso", pd.Series(dtype=float)), errors="coerce"
).fillna(0)

print(f"✅ Resultados finales: {len(df_final):,}")

# =========================================================
# 🔹 EXCEL FORMATEADO
# =========================================================

print("\n📁 Generando Excel...")

COLS_EXCEL = [c for c in [
    "origen",
    "nombre_entidad",
    "detalle_del_objeto_a_contratar",
    "modalidad_de_contratacion",
    "cuantia_proceso",
    "departamento_entidad",
    "municipio_entidad",
    "fecha_de_cargue_en_el_secop",
    "estado_del_proceso",
    "ruta_proceso_en_secop_i",
] if c in df_final.columns]

ANCHOS = {
    "origen": 12,
    "nombre_entidad": 35,
    "detalle_del_objeto_a_contratar": 60,
    "modalidad_de_contratacion": 25,
    "cuantia_proceso": 18,
    "departamento_entidad": 20,
    "municipio_entidad": 20,
    "fecha_de_cargue_en_el_secop": 22,
    "estado_del_proceso": 18,
    "ruta_proceso_en_secop_i": 18,
}

COLOR_HEADER    = "1F4E79"
COLOR_FILA_PAR  = "EEF4FF"
COLOR_FILA_IMPAR= "FFFFFF"

borde = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

excel_df = df_final[COLS_EXCEL].copy()
excel_df["cuantia_proceso"] = excel_df["cuantia_proceso"].apply(
    lambda x: f"${x:,.0f}" if pd.notnull(x) else ""
)

# Extraer URLs antes de limpiar
url_col = "ruta_proceso_en_secop_i"
urls = (
    excel_df[url_col].apply(extraer_url).tolist()
    if url_col in excel_df.columns else [""] * len(excel_df)
)
url_col_idx = COLS_EXCEL.index(url_col) + 1 if url_col in COLS_EXCEL else None

excel_df = excel_df.apply(lambda col: col.map(limpiar_valor))

wb = Workbook()
ws = wb.active
ws.title = "Alertas SECOP I"

# Cabecera
ws.append(COLS_EXCEL)
for cell in ws[1]:
    cell.font      = Font(bold=True, color="FFFFFF", size=11)
    cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = borde
ws.row_dimensions[1].height = 30

# Filas
for i, (row_data, url_val) in enumerate(
    zip(excel_df.itertuples(index=False), urls), start=2
):
    ws.append(list(row_data))
    color = COLOR_FILA_PAR if i % 2 == 0 else COLOR_FILA_IMPAR

    for j, cell in enumerate(ws[i], start=1):
        cell.fill      = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border    = borde

    if url_col_idx and url_val and str(url_val).startswith("http"):
        cell           = ws.cell(row=i, column=url_col_idx)
        cell.hyperlink = str(url_val)
        cell.value     = "🔗 Ver proceso"
        cell.font      = Font(color="0563C1", underline="single")

    ws.row_dimensions[i].height = 40

# Anchos de columna
for i, col_name in enumerate(COLS_EXCEL, start=1):
    ws.column_dimensions[get_column_letter(i)].width = ANCHOS.get(col_name, 20)

ws.freeze_panes = "A2"

excel_ok = False
try:
    wb.save(ARCHIVO_EXCEL)
    print(f"✅ Excel guardado: {ARCHIVO_EXCEL}")
    excel_ok = True
except Exception as e:
    print(f"❌ No se pudo guardar el Excel: {e}")

# =========================================================
# 🔹 HTML CORREO
# =========================================================

total       = len(df_final)
entidades   = df_final["nombre_entidad"].nunique() if "nombre_entidad" in df_final.columns else 0
valor_total = df_final["cuantia_proceso"].sum()
valor_prom  = df_final["cuantia_proceso"].mean()

top5 = df_final.sort_values(by="cuantia_proceso", ascending=False).head(5)
df_final["ruta_proceso_en_secop_i"] = df_final.get(
    "ruta_proceso_en_secop_i", pd.Series(dtype=str)
).apply(extraer_url)

filas_html = ""
for _, row in top5.iterrows():
    link      = extraer_url(row.get("ruta_proceso_en_secop_i", ""))
    link_html = f"<a href='{link}' style='color:#1F4E79;text-decoration:none;'>🔗 Ver</a>" if link else "—"
    cuantia   = safe_float(row.get("cuantia_proceso", 0))
    filas_html += f"""
    <tr>
        <td style="padding:8px;border-bottom:1px solid #e0e0e0;font-size:13px;color:#333;">
            {str(row.get('nombre_entidad',''))[:40]}
        </td>
        <td style="padding:8px;border-bottom:1px solid #e0e0e0;font-size:13px;color:#555;">
            {str(row.get('detalle_del_objeto_a_contratar',''))[:70]}...
        </td>
        <td style="padding:8px;border-bottom:1px solid #e0e0e0;font-size:13px;
                   color:#1a7f4b;font-weight:bold;">
            ${cuantia:,.0f}
        </td>
        <td style="padding:8px;border-bottom:1px solid #e0e0e0;font-size:13px;">
            {link_html}
        </td>
    </tr>"""

resumen_html = f"""
<html>
<body style="font-family:Arial,sans-serif;background:#f4f6f9;padding:20px;margin:0;">
<div style="max-width:700px;margin:auto;background:#ffffff;border-radius:10px;
            overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.1);">

  <div style="background:#1F4E79;padding:28px 32px;">
    <h1 style="color:#ffffff;margin:0;font-size:22px;">
      🎓 Oportunidades SECOP I para Universidades
    </h1>
    <p style="color:#a8c8e8;margin:6px 0 0;font-size:14px;">
      Contratos convocados · {inicio.strftime('%Y-%m-%d')} → {fecha_str} · {date.today().strftime("%d de %B de %Y")}
    </p>
  </div>

  <div style="display:flex;gap:0;border-bottom:3px solid #1F4E79;">
    <div style="flex:1;padding:20px;text-align:center;border-right:1px solid #e0e0e0;">
      <p style="margin:0;font-size:28px;font-weight:bold;color:#1F4E79;">{total}</p>
      <p style="margin:4px 0 0;font-size:12px;color:#888;">OPORTUNIDADES</p>
    </div>
    <div style="flex:1;padding:20px;text-align:center;border-right:1px solid #e0e0e0;">
      <p style="margin:0;font-size:28px;font-weight:bold;color:#1F4E79;">{entidades}</p>
      <p style="margin:4px 0 0;font-size:12px;color:#888;">ENTIDADES</p>
    </div>
    <div style="flex:1;padding:20px;text-align:center;border-right:1px solid #e0e0e0;">
      <p style="margin:0;font-size:22px;font-weight:bold;color:#1a7f4b;">${valor_total:,.0f}</p>
      <p style="margin:4px 0 0;font-size:12px;color:#888;">VALOR TOTAL</p>
    </div>
    <div style="flex:1;padding:20px;text-align:center;">
      <p style="margin:0;font-size:22px;font-weight:bold;color:#1a7f4b;">${valor_prom:,.0f}</p>
      <p style="margin:4px 0 0;font-size:12px;color:#888;">VALOR PROMEDIO</p>
    </div>
  </div>

  <div style="padding:24px 32px;">
    <h2 style="font-size:16px;color:#1F4E79;margin:0 0 16px;">💰 Top 5 Contratos por Valor</h2>
    <table style="width:100%;border-collapse:collapse;">
      <thead>
        <tr style="background:#f0f4f8;">
          <th style="padding:10px 8px;text-align:left;font-size:12px;color:#555;">ENTIDAD</th>
          <th style="padding:10px 8px;text-align:left;font-size:12px;color:#555;">DESCRIPCIÓN</th>
          <th style="padding:10px 8px;text-align:left;font-size:12px;color:#555;">VALOR</th>
          <th style="padding:10px 8px;text-align:left;font-size:12px;color:#555;">LINK</th>
        </tr>
      </thead>
      <tbody>{filas_html}</tbody>
    </table>
  </div>

  <div style="background:#f0f4f8;padding:16px 32px;text-align:center;">
    <p style="margin:0;font-size:12px;color:#999;">
      📎 El archivo Excel completo se adjunta a este correo
    </p>
  </div>

</div>
</body>
</html>
"""

# =========================================================
# 🔹 ENVÍO CORREO
# =========================================================

print("\n📧 Enviando correo...")

msg = MIMEMultipart("alternative")
msg["Subject"] = f"🎓 SECOP I Universidades · {fecha_str} ({total} oportunidades)"
msg["From"]    = EMAIL_SENDER
msg["To"]      = ", ".join(EMAIL_TO)
msg.attach(MIMEText(resumen_html, "html"))

if excel_ok:
    try:
        with open(ARCHIVO_EXCEL, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename={ARCHIVO_EXCEL}")
        msg.attach(part)
    except Exception as e:
        print(f"⚠️  No se pudo adjuntar Excel: {e}")

try:
    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.starttls()
        server.login(EMAIL_SENDER, EMAIL_PASSWORD)
        server.sendmail(EMAIL_SENDER, EMAIL_TO, msg.as_string())
    print("✅ Correo enviado correctamente")
except Exception as e:
    print(f"❌ Error al enviar correo: {e}")

print("\n🏁 Pipeline finalizado")
